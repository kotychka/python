import pandas as pd
import requests
from bs4 import BeautifulSoup
# from openpyxl import Workbook
from datetime import date, timedelta

def append_df_to_excel(filename, df, sheet_name='Sheet1', startrow=None,
                       truncate_sheet=False,
                       **to_excel_kwargs):
    """
    Append a DataFrame [df] to existing Excel file [filename]
    into [sheet_name] Sheet.
    If [filename] doesn't exist, then this function will create it.
  
    Parameters:
      filename : File path or existing ExcelWriter
                 (Example: '/path/to/file.xlsx')
      df : dataframe to save to workbook
      sheet_name : Name of sheet which will contain DataFrame.
                   (default: 'Sheet1')
      startrow : upper left cell row to dump data frame.
                 Per default (startrow=None) calculate the last row
                 in the existing DF and write to the next row...
      truncate_sheet : truncate (remove and recreate) [sheet_name]
                       before writing DataFrame to Excel file
      to_excel_kwargs : arguments which will be passed to `DataFrame.to_excel()`
                        [can be dictionary]
  
    Returns: None
    """
    from openpyxl import load_workbook
  
    # ignore [engine] parameter if it was passed
    if 'engine' in to_excel_kwargs:
        to_excel_kwargs.pop('engine')
  
    writer = pd.ExcelWriter(filename, engine='openpyxl')
  
    try:
        # try to open an existing workbook
        writer.book = load_workbook(filename)
  
        # get the last row in the existing Excel sheet
        # if it was not specified explicitly
        if startrow is None and sheet_name in writer.book.sheetnames:
            startrow = writer.book[sheet_name].max_row
  
        # truncate sheet
        if truncate_sheet and sheet_name in writer.book.sheetnames:
            # index of [sheet_name] sheet
            idx = writer.book.sheetnames.index(sheet_name)
            # remove [sheet_name]
            writer.book.remove(writer.book.worksheets[idx])
            # create an empty sheet [sheet_name] using old index
            writer.book.create_sheet(sheet_name, idx)
  
        # copy existing sheets
        writer.sheets = {ws.title:ws for ws in writer.book.worksheets}
    except FileNotFoundError:
        # file does not exist yet, we will create it
        pass
  
    if startrow is None:
        startrow = 0
  
    # write out the new sheet
    df.to_excel(writer, sheet_name, startrow=startrow, **to_excel_kwargs)

    # write date

  
    # save the workbook
    writer.save()

# Get dates
today = date.today().strftime('%d.%m.%Y')
yest = (date.today() - timedelta(days=1)).strftime('%d.%m.%Y')

# 
page = requests.get('https://www.worldometers.info/coronavirus/?')
soup = BeautifulSoup(page.content, 'html.parser')

# Take div with all needed data
table_today = soup.find(id='main_table_countries_today')
table_yest = soup.find(id='main_table_countries_yesterday')

items_today = table_today.find_all(class_='total_row_body')
items_yest = table_yest.find_all(class_='total_row_body')

# Take data to variables
header = str(table_today.find('thead').get_text().replace(',', '').replace('\\n', ' ').split('\\n')).replace('/\\n', '-').replace('/', '-').replace('\\xa0', '')[6:-19].split('\\n')

countries_today = str([item.find(class_='total_row').get_text() for item in items_today])[12:-9].split('\\n')
countries_today.insert(0, today)
countries_yest = str([item.find(class_='total_row').get_text() for item in items_yest])[12:-9].split('\\n')
countries_yest.insert(0, yest)

# Data frame
countries_data = pd.DataFrame([countries_yest,countries_today], columns=None)	

# Write data in existing file
append_df_to_excel('d:/Cloud/Nastya/IT/python/projects/covid/CoVid_data.xlsx', countries_data, sheet_name='Data', header=False, index=False)
